#! /usr/bin/env python3

# Authors: Victor Castillo, Pau Treig i Joan Jara
# Data: 05/03/2019
# Version: 1
# Parameters: Cap parametre
# Description: Script en python que llegeix els fitxers excel dels kahoot que hi ha en un directori per tal de anonimitzar els noms dels participants i guarda en un fitxer txt la traduccio del nom anonim al real. Recalcula els punts en cas dhaverhi una modificacio a lexcel i n'extreu un ranquing.

#importem tots els moduls necessaris
import os
import shutil
from glob import glob 
import sys
import openpyxl

#classe amb els atributs que necessitarem per a cada participant, amb getters i setters per a cada atribut
class participant:
	def __init__(self, name, puntuacio, nameAnonim, respostesC, respostesI):
		self.name=name
		self.puntuacio=puntuacio
		self.nameAnonim=nameAnonim
		self.respostesC=respostesC
		self.respostesI=respostesI
	def __str__(self):
		return "Nom: "+self.name+"\nPuntuacio: "+str(self.puntuacio)+"\nAnonim: "+self.nameAnonim+"\nRespostesCorrectes: "+str(self.respostesC)+"\nRespostesIncorrectes: "+str(self.respostesI)+"\n"
	def getNom(self):
		return self.name
	def getPuntuacio(self):
		return self.puntuacio
	def getNomA(self):
		return self.nameAnonim
	def getRespostesC(self):
		return self.respostesC
	def getRespostesI(self):
		return self.respostesI
	def setPuntuacio(self, puntuacio2):
		self.puntuacio=self.getPuntuacio()+puntuacio2
	def setRespostesC(self, respostesC2):
		self.respostesC=self.getRespostesC()+respostesC2
	def setRespostesI(self, respostesI2):
		self.respostesI=self.getRespostesI()+respostesI2

#funcio que passat un directori per parametre, copia els fitxers excels trobats en el directori en excels nous.
def copiaExcels(directori):
	for fitxer in glob(directori+"*.xlsx"):
		fitxer_copi=directori+"nou_"+fitxer.split("/")[-1]
		with open(fitxer, 'rb') as forigen:
			with open (fitxer_copi, 'wb') as fdestino:
				shutil.copyfileobj(forigen,fdestino)
				
#funcio que passat un directori per parametre, anonimitza tots els noms de tots els fitxers excel que trobi al directori, i escriu en un fitxer les relacions "Nom Anonim - Nom Real"
def fitxerAnonim(directori):
	nomTotal=[]    #llista de noms sense repetir on es guarden tots els noms de tots els fitxers excel
	diccionariNoms={}    #diccionari on es passa per clau un nom real i retorna el nom anonim
	diccionariAnonims={}    #diccionari on es passa per clau un nom anonim i retorna el nom real
	print("Anonimitzant...")
	anonim=1    #aquest sera lindex del numero danonims que tenim
	fitxerNoms=open("fitxer_noms.txt","w")
	dash='-'*40
	fitxerNoms.write(dash+"\n")
	fitxerNoms.write('{0:10}\t\t{1:9}'.format("NOM ANONIM","NOM REAL\n"))    #la funcio format ens ajudara al llarg del codi a escriure de forma ben tabulada als fitxers de text
	fitxerNoms.write(dash+"\n")
	for fitxer in glob(directori+"nou_*.xlsx"):    #glob ens retorna la llista de fitxers que acaben en '.xlsx' dins del directori, i recorrem aquests fitxers amb el for
		llistaNomsRepetits=[]    #llista de noms sense repetir en un mateix fitxer excel, en cas de que hi hagi un nom repetit dins dun fitxer lanularem
		excel = openpyxl.load_workbook(fitxer)
		overview = excel['Overview']
		nJugadors=overview['B4'].value
		nJugadors=nJugadors.split()[0]    #agafem el numero de jugadors a partir de la cella B4 de la fulla Overview, a la que li retallem l'string "players" i ens quedem amb el numero
		finalScores = excel['Final Scores']
		questionSummary = excel['Question Summary']
		fullesQuestions=excel.sheetnames[3:-1]    #obtenim la llista amb les fulles corresponents a cada questio
		index=4    #inicialitzem lindex a 4 ja que els noms comencen a la fila 4 de la fulla Final Scores
		while(index<int(nJugadors)+4):    #recorrem els n jugadors del fitxer excel, sortira del bucle quan index sigui mes gran o igual que el numero de jugadors
			nom = finalScores['B'+str(index)].value    #agafem el nom a partir de recorrer la columna B de la fulla Final Scores
			if nom not in nomTotal:    #només treballarem amb els noms que no estiguin repetits
				nomTotal.append(nom)
				fitxerNoms.write('{0:10}\t\t{1:10}'.format("Anonim"+str(anonim),nom+"\n"))    #escrivim la relacio al fitxer de text 'fitxer_noms.txt'
				diccionariNoms.update({nom:'Anonim'+str(anonim)})
				diccionariAnonims.update({'Anonim'+str(anonim):nom})    #actualitzem ambdos diccionaris
				anonim=anonim+1
			if nom not in llistaNomsRepetits:   #si el nom no esta repetit en el fitxer escriurem el nom anonim corresponent a les fulles Final Score i Question Summary
				llistaNomsRepetits.append(nom)
				finalScores['B'+str(index)]=diccionariNoms.get(nom)
				name = questionSummary['B'+str(index)].value
				questionSummary['B'+str(index)]=diccionariNoms.get(name)
			else:    #si el nom esta repetit dins dun mateix fitxer excel, lanularem (a partir del segon, es a dir, el primer el deixarem acceptat)
				finalScores['B'+str(index)]='ANULAT(Nom Repetit)'
				questionSummary['B'+str(index)]='ANULAT(Nom Repetit)'
			index=index+1
		
		for question in fullesQuestions:    #bucle que recorre les fulles de les questions i escriu el nom anonim corresponent a la cella
			question = excel[question]
			index=15
			while(index<int(nJugadors)+15):
				name=question['A'+str(index)].value
				question['A'+str(index)]=diccionariNoms.get(name)
				index=index+1
				
		index=2
		totalPreguntesJugador=len(fullesQuestions)*int(nJugadors)    #calculem el numero de celles que haurem de modificar a partir de multiplicar el numero de questions pel numero de jugadors
		rawQuestion = excel['RawReportData Data']
		while(index<totalPreguntesJugador+2):    #bucle que recorre la fulla 'RawReportData Data' i escriu el nom anonim corresponent a cada cella de la columna I
			name=rawQuestion['I'+str(index)].value
			rawQuestion['I'+str(index)]=diccionariNoms.get(name)
			index=index+1
		excel.save(fitxer)    #guardem els canvis
	fitxerNoms.close()
	print("fitxer_noms.txt creat")
	return diccionariAnonims
	
def recalcular(directori):    #funcio que recalcula els punts de cada jugador i, en cas de que el resultat sigui diferent al de la fulla Final Scores, ho notifica en el fitxer de text 'actualitzacio_punts.txt'
	fitxerPunts=open("actualitzacio_punts.txt","w")
	dash='-'*70
	print("Recalculant puntuacions...")
	for fitxer in glob(directori+"*nou_*.xlsx"):
		fitxerPunts.write(fitxer+"\n")
		fitxerPunts.write(dash+"\n")
		fitxerPunts.write('{0:10}\t\t{1:15}\t\t{2:15}\n'.format("NOM","PUNTS_ORIGINALS", "PUNTS_CALCULATS"))   #capcalera del fitxer
		fitxerPunts.write(dash+"\n")
		excel = openpyxl.load_workbook(fitxer)
		finalScores = excel['Final Scores']
		questionSummary = excel['Question Summary']
		overview = excel['Overview']
		numJugadors=overview['B4'].value
		numJugadors=numJugadors.split()[0]
		numQuestions=len(excel.sheetnames[3:-1])    #treiem el numero de questions a partir daplicar la funcio len a la llista amb les fulles corresponents a les questions
		
		fila=4
		while(fila<int(numJugadors)+4):
			columna=4    #inicialitzem la fila i la columna a 4, que es on comencen els valors que ens interessen de la fulla Question Summary
			numCorrectes=0
			punts=0
			while(columna<(numQuestions*2+4)):
				valor=int(questionSummary.cell(row=fila, column=columna).value)    #obtenim el valor de la cella
				if(valor>0):    #si el valor es mes gran que 0, vol dir que la resposta es correcta
					numCorrectes+=1
					punts+=valor
				columna+=2    #augmentem la columna de dos en dos ja que les columnes amb els valors estan separades per una columna amb la resposta que el jugador ha triat
			valorOriginal=finalScores.cell(row=fila, column=3).value
			if(punts != int(valorOriginal)):    #un cop ja hem sumat els punts de totes les preguntes, comparem la suma amb el valor original. Si es diferent vol dir que hi ha hagut alguna modificacio i per tant actualitzarem els valors de la fulla 'Final Scores' i ho escriurem al fitxer de text.
				finalScores.cell(row=fila, column=3).value=punts
				finalScores.cell(row=fila, column=4).value=numCorrectes
				finalScores.cell(row=fila, column=5).value=numQuestions-numCorrectes
				fitxerPunts.write('{0:10}\t{1:15d}\t\t{2:15d}\n'.format(finalScores.cell(row=fila, column=2).value,int(valorOriginal),int(punts)))
			fila+=1
		excel.save(fitxer)    #guardem els canvis
		fitxerPunts.write("\n\n")
	print("actualitzacio_punts.txt creat")
	fitxerPunts.close()

def ordenarRanquing(taula, esquerra, dreta):    #funcio recursiva que ordena el ranquing a partir de la puntuacio (quicksort)
	i=esquerra
	j=dreta
	pivot=(esquerra + dreta)/2
	x=taula[int(pivot)].getPuntuacio()
	while( i <= j ):
		while taula[i].getPuntuacio()<x and j<=dreta:
			i=i+1
		while x<taula[j].getPuntuacio() and j>esquerra:
			j=j-1
		if i<=j:
			aux = taula[i]; taula[i] = taula[j]; taula[j] = aux;
			i=i+1;  j=j-1;
		if esquerra < j:
			ordenarRanquing( taula, esquerra, j );
	if i < dreta:
		ordenarRanquing( taula, i, dreta );
        
def ranquing(directori, diccionari):    #funcio que crea un ranquing al fitxer de text 'ranquing.txt' en base a les dades dels fitxers excel
	print("Creant el ranquing...")
	fitxerRanquing=open("ranquing.txt","w")
	dash='-'*105
	fitxerRanquing.write(dash+"\n")
	fitxerRanquing.write('{0:25}\t{1:10}\t   {2:10}\t\t{3:15}\t{4:15}\n'.format("NOM REAL","NOM ANONIM","PUNTUACIO","RES_CORRECTES","RES_INCORRECTES"))    #capcalera del fitxer
	fitxerRanquing.write(dash+"\n")
	taula_ranquing=[]    #llista amb tots els participants
	for fitxer in glob(directori+"nou_*.xlsx"):
		excel = openpyxl.load_workbook(fitxer)
		finalScores = excel['Final Scores']
		overview = excel['Overview']
		numJugadors=overview['B4'].value
		numJugadors=numJugadors.split()[0]
		index=4
		while(index<int(numJugadors)+4):
			nom=finalScores['B'+str(index)].value
			if(nom!='ANULAT(Nom Repetit)'):    #si el participant ha estat anulat no participa al ranquing
				alumne= participant(diccionari.get(nom),finalScores['C'+str(index)].value,nom,finalScores['D'+str(index)].value,finalScores['E'+str(index)].value)    #creem un objecte participant amb les dades necessaries (Nom Real, Punts, Nom Anonim, Respostes Correctes, Respostes Incorrectes)
				trobat=False
				indexT=0
				while(trobat==False and indexT<len(taula_ranquing)):    #recorrem la llista amb tots els participants per a saber si lalumne actual ja forma part o no de la llista
					if alumne.getNom()==taula_ranquing[indexT].getNom():
						trobat=True    #en cas de que ja formi part del ranquing, trobat=True
					indexT+=1
				if(trobat==False):    #si lalumne actual no formava part del ranquing simplement se lafegeix
					taula_ranquing.append(alumne)
				else:    #si ja formava part del ranquing, vol dir que li haurem de sumar els resultats actuals als que ja tenia
					taula_ranquing[indexT-1].setPuntuacio(alumne.getPuntuacio())    #aixo ho farem mitjancant aquests 3 setters, que sumen el valor que ja teniem amb el passat per parametre
					taula_ranquing[indexT-1].setRespostesC(alumne.getRespostesC())
					taula_ranquing[indexT-1].setRespostesI(alumne.getRespostesI())
			index+=1
	ordenarRanquing(taula_ranquing,0, len(taula_ranquing)-1)   #ordenem el ranquing
	for alumne in reversed(taula_ranquing):    #escrivim el resultat al fitxer de text
		fitxerRanquing.write('{0:25}\t{1:10}\t{2:10d}\t{3:15d}\t{4:15d}\n'.format(alumne.getNom(),alumne.getNomA(),alumne.getPuntuacio(),alumne.getRespostesC(),alumne.getRespostesI()))
	fitxerRanquing.close()
	print("ranquing.txt creat")
#MAIN
print("En quin directori et vols posicionar per a fer el tractament de fulls excel?")
directori=input()
if directori[-1]!="/":    #si lusuari introdueix el directori sense la barra final, se li afegeix
			directori=directori+"/"
			
if(os.path.isdir(directori)):    #si el directori introduit existeix, cridem a les tres funcions implementades
	copiaExcels(directori)
	diccionari=fitxerAnonim(directori)
	recalcular(directori)
	ranquing(directori,diccionari)
else:
	print("Error: Has de posar el path d'un directori valid i existent")
